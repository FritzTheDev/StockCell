#import the api wrapper, date-time, json, and an excel interface module
from alpha_vantage.timeseries import TimeSeries
import datetime
import json
from openpyxl import Workbook
from decimal import Decimal
# Functional programming approach below: Just using it as notes.
# date = None
# workingRow = 2
# introMessage = 'Welcome to StockCell,'
# tickerInput =  'Please input your desired ticker   > '
# tickerError = 'Not a valid ticker'
# dateInput = 'What date would you like data for? > '
# dateError = 'Please input date in the YYYY-MM-DD format'
# def api_call():
#     while True:
#         try:
#             global ticker
#             ticker = input(tickerInput)
#             jsondict, meta = ts.get_daily(ticker)
#             return jsondict
#             break
#         except ValueError:
#             print(errorMessage)
# def date_entry(): # TODO: Regex for date entry validation
#     while True:
#         try:
#             global date
#             date = (input(dateInput))
#             return date
#         except ValueError:
#             print(dateError)
# def data_reader(data, date):
#     open_value = data[date]['1. open']
#     high_value = data[date]['2. high']
#     low_value = data[date]['3. low']
#     close_value = data[date]['4. close']
#     vol_value = data[date]['5. volume']
#     value_dict = (open_value, high_value, low_value, close_value, vol_value)
#     return value_dict
# def rowfill(fivevalues):
#     global workingRow
#     #TODO: Add date column
#     ws['B'+ str(workingRow)] = ticker.upper()
#     ws['C'+ str(workingRow)] = fivevalues[0]
#     ws['D'+ str(workingRow)] = fivevalues[1]
#     ws['E'+ str(workingRow)] = fivevalues[2]
#     ws['F'+ str(workingRow)] = fivevalues[3]
#     ws['G'+ str(workingRow)] = fivevalues[4]
#     workingRow += 1
# ticker = None
# print(introMessage)
# fivevalues = data_reader(api_call(), date_entry())
# wb = Workbook()
# ws = wb.active
# ws['A1'] = 'Date'
# ws['B1'] = 'Ticker'
# ws['C1'] = 'Open'
# ws['D1'] = 'High'
# ws['E1'] = 'Low'
# ws['F1'] = 'Close'
# ws['G1'] = 'Volume'
# rowfill(fivevalues)
# rowfill(fivevalues)
# wb.save("tester.xlsx")
ts = TimeSeries('N1RDW0MA9WZPJGCK')
class FillRow():
    tickerError = 'Invalid Ticker, Please Try Again'
    def __init__(self, row):
        def api_call(ticker):
            while True:
                try:
                    data, meta = ts.get_daily(ticker)
                    return data
                    break
                except ValueError:
                    print(tickerError)
        tickerGet = input('Ticker: > ')
        data = api_call(tickerGet)
        date = input('Date: > ')
        self.row = row
        self.open_value = data[date]['1. open']
        self.high_value = data[date]['2. high']
        self.low_value = data[date]['3. low']
        self.close_value = data[date]['4. close']
        self.vol_value = data[date]['5. volume']
        ws['A' + str(row)] = date
        ws['B' + str(row)] = tickerGet
        ws['C' + str(row)] = self.open_value
        ws['D' + str(row)] = self.high_value
        ws['E' + str(row)] = self.low_value
        ws['F' + str(row)] = self.close_value
        ws['G' + str(row)] = self.vol_value
        wb.save('tester.xlsx')
wb = Workbook()
ws = wb.active
def wb_setup():
    ws['A1'] = 'Date'
    ws['B1'] = 'Ticker'
    ws['C1'] = 'Open'
    ws['D1'] = 'High'
    ws['E1'] = 'Low'
    ws['F1'] = 'Close'
    ws['G1'] = 'Volume'
wb_setup()
def main():
    row = 2
    # keepGoing = input('Would you like to add another stock? (y/n)')
    keepGoing = 'y'
    while keepGoing == 'y':
        FillRow(row)
        row += 1
        keepGoing = input('Would you like to add another stock? (y/n)')
main()
wb.save()
